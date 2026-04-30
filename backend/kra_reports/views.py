from django.shortcuts import render
import io
from django.http import HttpResponse
# Create your views here.
from django.db import transaction
from django.shortcuts import get_object_or_404

from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from kra_cycle.models import (
    Employee,
    KRACycle,
    KRACycleStage,
    EmployeeKRACycle,
    EmployeeKRACycleCategory,
    EmployeeKRALevel,
    KRALevel,
    KRA,
    KRACategory,
    Stage,
    Level,
    Rating,
    AuditLog,
)

HR_ROLES      = {"Admin" , "HR" , "Vertical Lead"}
LEAD_ROLES    = {"Manager" , "Team Lead"}
EMPLOYEE_ROLE = "Employee"


def _get_caller(request):
    return request.user


def _is_hr(employee):
    # Check bridge table first, fall back to direct role FK
    if employee.employee_roles.filter(role__name__in=HR_ROLES).exists():
        return True
    return bool(employee.role and employee.role.name in HR_ROLES)


def _is_lead(employee):
    # Check bridge table first, fall back to direct role FK
    if employee.employee_roles.filter(role__name__in=LEAD_ROLES).exists():
        return True
    return bool(employee.role and employee.role.name in LEAD_ROLES)


def _caller_can_act_on(caller, target_employee_id):
    if _is_hr(caller):
        return True
    return Employee.objects.filter(id=target_employee_id, manager_id=caller.id).exists()


def _audit(request, action, entity, entity_id, old_data=None, new_data=None):
    AuditLog.objects.create(
        employee = _get_caller(request),
        action = action,
        entity = entity,
        entity_id = entity_id,
        old_data = old_data,
        new_data = new_data,
        ip_address = request.META.get("REMOTE_ADDR"),
    )
    
class ReportView(APIView):
    """
    GET /api/v1/kra/cycles/{cycle_id}/report?employee_id=&department_id=&category_id=
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, cycle_id):
        caller               = _get_caller(request)
        employee_id_filter   = request.query_params.get('employee_id')
        department_id_filter = request.query_params.get('department_id')
        category_id_filter   = request.query_params.get('category_id')

        cycle = get_object_or_404(KRACycle, id=cycle_id, is_deleted=False)

        ekc_qs = EmployeeKRACycle.objects.filter(
            kra_cycle_id=cycle_id
        ).select_related('employee__department', 'employee__level')

        if not _is_hr(caller):
            ekc_qs = ekc_qs.filter(employee__manager_id=caller.id)
        if employee_id_filter:
            ekc_qs = ekc_qs.filter(employee_id=employee_id_filter)
        if department_id_filter:
            ekc_qs = ekc_qs.filter(employee__department_id=department_id_filter)

        report_employees = []

        for ekc in ekc_qs:
            kra_rows_qs = ekc.kra_level_rows.select_related(
                'kra_level__kra__category',
                'self_rating',
                'lead_rating',
            )
            if category_id_filter:
                kra_rows_qs = kra_rows_qs.filter(
                    kra_level__category_id=category_id_filter
                )

            kra_rows = list(kra_rows_qs)

            # Compute averages (skip nulls)
            self_ratings  = [r.self_rating.rating for r in kra_rows if r.self_rating]
            lead_ratings  = [r.lead_rating.rating for r in kra_rows if r.lead_rating]
            self_avg      = round(sum(self_ratings) / len(self_ratings), 2) if self_ratings else None
            lead_avg      = round(sum(lead_ratings) / len(lead_ratings), 2) if lead_ratings else None

            # Group by category
            cat_map = {}
            for r in kra_rows:
                kl       = r.kra_level
                cat_id   = getattr(kl, 'category_id', None)
                cat_name = getattr(getattr(kl, 'category', None), 'name', None)
                key      = cat_id

                if key not in cat_map:
                    # Fetch weightage for this category
                    wobj = ekc.categories.filter(category_id=cat_id).first()
                    cat_map[key] = {
                        'category_id':   cat_id,
                        'category_name': cat_name,
                        'weightage':     wobj.weightage if wobj else None,
                        'kras':          [],
                    }
                cat_map[key]['kras'].append({
                    'employee_kra_level_id': r.id,
                    'kra_name':              getattr(kl, 'name', None),
                    'self_rating':           r.self_rating.rating if r.self_rating else None,
                    'lead_rating':           r.lead_rating.rating if r.lead_rating else None,
                    'self_comment':          r.self_comment,
                    'lead_comment':          r.lead_comment,
                })

            e = ekc.employee
            report_employees.append({
                'employee_id':            e.id,
                'full_name':              f'{e.first_name} {e.last_name}',
                'department':             e.department.department_name if e.department else None,
                'level':                  e.level.name if e.level else None,
                'overall_self_rating_avg': self_avg,
                'overall_lead_rating_avg': lead_avg,
                'categories':             list(cat_map.values()),
            })
        #  AUDIT LOG
        _audit(
            request,
            "REPORT_VIEWED",
            "KRACycle",
            cycle.id,
            new_data={
                "employee_filter": employee_id_filter,
                "department_filter": department_id_filter,
                "category_filter": category_id_filter,
                "records_returned": len(report_employees),
                "viewer_role": caller.role.name if getattr(caller, "role", None) else None,
            }
        )

        return Response({
            'cycle_id':   cycle.id,
            'cycle_name': cycle.name,
            'employees':  report_employees,
        }, status=status.HTTP_200_OK)



# 20. Report – Export (Excel / PDF)                                       


class ReportExportView(APIView):
    """
    GET /api/v1/kra/cycles/{cycle_id}/report/export?format=excel|pdf
    Returns a downloadable Excel file (openpyxl).
    PDF stub included — wire up weasyprint/reportlab when ready.

    Install: pip install openpyxl
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, cycle_id):
        export_format        = request.query_params.get('file_type', '').lower()
        employee_id_filter   = request.query_params.get('employee_id')
        department_id_filter = request.query_params.get('department_id')

        if export_format not in ('excel', 'pdf'):
            return Response(
                'Invalid format. Use excel or pdf',
                status=status.HTTP_400_BAD_REQUEST,
            )

        caller = _get_caller(request)
        cycle  = get_object_or_404(KRACycle, id=cycle_id, is_deleted=False)

        ekc_qs = EmployeeKRACycle.objects.filter(
            kra_cycle_id=cycle_id
        ).select_related('employee__department', 'employee__level')

        if not _is_hr(caller):
            ekc_qs = ekc_qs.filter(employee__manager_id=caller.id)
        if employee_id_filter:
            ekc_qs = ekc_qs.filter(employee_id=employee_id_filter)
        if department_id_filter:
            ekc_qs = ekc_qs.filter(employee__department_id=department_id_filter)

        if export_format == 'excel':
            file_bytes = self._build_excel(cycle, ekc_qs)
            content_type = (
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'kra_report_cycle_{cycle_id}.xlsx'
        else:
            # PDF stub — replace with weasyprint/reportlab
            file_bytes = b'%PDF stub - implement with weasyprint or reportlab'
            content_type = 'application/pdf'
            filename = f'kra_report_cycle_{cycle_id}.pdf'
            
        #  AUDIT LOG
        _audit(
            request,
            "REPORT_EXPORTED",
            "KRACycle",
            cycle.id,
            new_data={
                "format": export_format,
                "employee_filter": employee_id_filter,
                "department_filter": department_id_filter,
                "records_count": ekc_qs.count(),
                "viewer_role": caller.role.name if getattr(caller, "role", None) else None,
            }
        )

        response = HttpResponse(file_bytes, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # Excel builder 
    @staticmethod
    def _build_excel(cycle, ekc_qs):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError('pip install openpyxl')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Performance Report'

        #  Header row 
        header_fill = PatternFill('solid', fgColor='1F4E79')
        header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        headers = [
            'Employee ID', 'Full Name', 'Department', 'Level',
            'KRA Level ID', 'KRA Name',
            'Self Rating', 'Self Comment',
            'Lead Rating', 'Lead Comment',
            'Progress Notes', 'Lead Progress Notes',
            'Description by Lead',
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = cell.font = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows 
        alt_fill = PatternFill('solid', fgColor='D6E4F0')
        row_idx  = 2

        for ekc in ekc_qs:
            e        = ekc.employee
            kra_rows = ekc.kra_level_rows.select_related(
                'kra_level', 'self_rating', 'lead_rating'
            )

            for r in kra_rows:
                kl = r.kra_level
                row_data = [
                    e.id,
                    f'{e.first_name} {e.last_name}',
                    e.department.department_name if e.department else '',
                    e.level.name if e.level else '',
                    kl.id if kl else '',
                    getattr(kl, 'name', '') or '',
                    r.self_rating.rating if r.self_rating else '',
                    r.self_comment or '',
                    r.lead_rating.rating if r.lead_rating else '',
                    r.lead_comment or '',
                    r.progress_notes or '',
                    r.lead_progress_notes or '',
                    r.description_by_lead or '',
                ]
                ws.append(row_data)

                # Alternate row shading
                if row_idx % 2 == 0:
                    for cell in ws[row_idx]:
                        cell.fill = alt_fill
                row_idx += 1

        # Auto-size columns (approximate)
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    